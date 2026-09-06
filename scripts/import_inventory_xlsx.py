#!/usr/bin/env python3
"""Import inventory rows from Excel into confirmed inventory_entries."""
from __future__ import annotations

import sys
import uuid
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

# Allow running from /tmp with app on PYTHONPATH
sys.path.insert(0, "/opt/food_checking")

from app.catalog.products import lookup_product  # noqa: E402
from app.catalog.units import STORAGE_UNITS  # noqa: E402
from app.db.session import SessionLocal  # noqa: E402
from app.services.inventory import get_or_create_named_product, get_or_create_product  # noqa: E402
from app.db.models import InventoryEntry  # noqa: E402

MOSCOW = ZoneInfo("Europe/Moscow")
ENTRY_DATE = date(2026, 9, 5)

UNIT_ALIASES = {
    "штука": "шт",
    "штуки": "шт",
    "штук": "шт",
    "шт": "шт",
    "пачки": "пачка",
    "пачек": "пачка",
    "пачка": "пачка",
    "кг": "кг",
    "килограмм": "кг",
    "десяток": "десяток",
    "банка": "банка",
    "банки": "банка",
    "палка": "палка",
    "батон": "батон",
    "бутылка": "бутылка",
    "упаковка": "упаковка",
    "пучок": "пучок",
}


def normalize_unit(raw: object) -> str | None:
    if raw is None:
        return None
    key = str(raw).strip().lower()
    mapped = UNIT_ALIASES.get(key, key)
    if mapped in STORAGE_UNITS:
        return mapped
    return None


def load_rows(path: Path) -> list[tuple[str, Decimal, str | None]]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows: list[tuple[str, Decimal, str | None]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i == 1:
            continue
        name, qty, unit = (row + (None, None, None))[:3]
        if name is None or str(name).strip() == "":
            continue
        if qty is None:
            print(f"SKIP empty qty: {name!r}")
            continue
        try:
            quantity = Decimal(str(qty))
        except (InvalidOperation, ValueError):
            print(f"SKIP bad qty {qty!r} for {name!r}")
            continue
        rows.append((str(name).strip(), quantity, normalize_unit(unit)))
    return rows


def main() -> int:
    path = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/05_09_2026.xlsx")
    apply = "--apply" in sys.argv
    rows = load_rows(path)
    print(f"file={path} rows={len(rows)} apply={apply} entry_date={ENTRY_DATE}")

    planned: list[tuple[str, Decimal, str, str]] = []
    # name_display, qty, unit, source_note
    for name, qty, excel_unit in rows:
        catalog = lookup_product(name)
        if catalog:
            planned.append((catalog.name, qty, catalog.unit, f"catalog←{name!r}"))
            continue
        if excel_unit:
            planned.append((name, qty, excel_unit, f"new←{name!r}"))
        else:
            print(f"SKIP unknown product/unit: {name!r} unit={excel_unit!r}")

    for item in planned:
        print(f"  {item[0]!r:30} {item[1]:>8} {item[2]:<10} ({item[3]})")

    if not apply:
        print("Dry-run only. Re-run with --apply to write DB.")
        return 0

    batch_id = str(uuid.uuid4())
    now = datetime.now(MOSCOW)
    with SessionLocal() as db:
        for name, qty, unit, note in planned:
            catalog = lookup_product(name)
            if catalog and note.startswith("catalog"):
                product = get_or_create_product(db, catalog)
                unit = product.unit
            else:
                product = get_or_create_named_product(db, name, unit)
            db.add(
                InventoryEntry(
                    product_id=product.id,
                    quantity=qty,
                    unit=unit,
                    status="confirmed",
                    source="excel",
                    transcript=f"import 05_09_2026.xlsx: {name}",
                    telegram_message_id=None,
                    batch_id=batch_id,
                    recorded_at=now,
                    entry_date=ENTRY_DATE,
                    confirmed_at=now,
                )
            )
        db.commit()
    print(f"OK wrote {len(planned)} confirmed inventory_entries batch_id={batch_id}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
