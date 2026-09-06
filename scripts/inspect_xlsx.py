#!/usr/bin/env python3
"""Inspect and optionally import inventory Excel into food_checking DB."""
from __future__ import annotations

import sys
from pathlib import Path

path = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/05_09_2026.xlsx")

try:
    import openpyxl
except ImportError:
    import subprocess

    subprocess.check_call(
        [sys.executable, "-m", "pip", "install", "-q", "openpyxl"],
    )
    import openpyxl

wb = openpyxl.load_workbook(path, data_only=True)
print("sheets:", wb.sheetnames)
for name in wb.sheetnames:
    ws = wb[name]
    print(f"\n=== sheet {name!r} dims={ws.dimensions} ===")
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > 40:
            print("... truncated ...")
            break
        print(i, row)
